from tkinter import *
from tkinter import ttk
from random import randint
from time import sleep
import serial
from openpyxl import Workbook, load_workbook

wb = load_workbook('productos.xlsx')
ws1 = wb['Sheet1']

root = Tk()
root.title("Ensayos")

estilo = ttk.Style()
estilo.theme_use('default')
estilo.configure("green.Horizontal.TProgressbar", foreground='green', background='green', thickness=70)

root.geometry('900x300')

def AgregarProducto(data1, data2, data3, data4, data5, data6, data7, data8, data9):
    x = 2
    data = []
    data.extend([data1,data2,data3,data4,data5,data6,data7,data8,data9])

    for row in range(1,1000):
        if(ws1.cell(row,1).value is None):
            ws1.cell(row=row,column=1).value = (ws1.cell(row-1,1).value) + 1
            for i in data:
                ws1.cell(row=row,column=x).value = i
                x = x + 1
            break

def Presionado(btn):
    if btn['bg'] == botonoff:
        btn.config(bg=botonon)
    else:
        btn.config(bg=botonoff)

def CambioClase(event, ):
    if clase.get() == 'Clase 2':
        boton4.config(state='disabled', bg=botonoff)
        boton5.config(state='disabled', bg=botonoff)
    else:
        boton4.config(state='normal')
        boton5.config(state='normal')

def VentanaAggProducto():
    data = []
    AggProducto = Toplevel(root)
    AggProducto.title("Agregar producto")
    AggProducto.geometry("290x410")
    Marca = Entry(AggProducto)
    Modelo = Entry(AggProducto)
    Descripcion = Entry(AggProducto)
    MonoTri = ttk.Combobox(AggProducto, state='readonly', width=18)
    MonoTri['values'] = ('Monofasico', 'Trifasico')
    PotenciaN = Entry(AggProducto)
    Tension = Entry(AggProducto)
    CorrienteM = Entry(AggProducto)
    Frecuencia = ttk.Combobox(AggProducto, state='readonly', width=18)
    Frecuencia['values'] = ('50Hz','60Hz', '50Hz/60Hz')
    Clase = ttk.Combobox(AggProducto, state='readonly', width=18)
    Clase['values'] = ('Clase 1', 'Clase 2')

    Marcalbl = Label(AggProducto, text='Marca:', width=11, anchor='w', pady=10).grid(row=1, column=0)
    Modelolbl = Label(AggProducto, text='Modelo:', width=11, anchor='w', pady=10).grid(row=2, column=0)
    Descripcionlbl = Label(AggProducto, text='Descripcion:', width=11, anchor='w', pady=10).grid(row=3, column=0)
    MonoTrilbl = Label(AggProducto, text='Fases:', width=11, anchor='w', pady=10).grid(row=4, column=0)
    PotenciaNlbl = Label(AggProducto, text='Potencia Nom.:', width=11, anchor='w', pady=10).grid(row=5, column=0)
    Watt = Label(AggProducto, text='W', width=2,  anchor='w').grid(row=5, column=2)
    Tensionlbl = Label(AggProducto, text='Tension:', width=11, anchor='w', pady=10).grid(row=6, column=0)
    Volt = Label(AggProducto, text='V', width=2,  anchor='w').grid(row=6, column=2)
    CorrienteMlbl = Label(AggProducto, text='Corriente Max.:', width=11, anchor='w', pady=10).grid(row=7, column=0)
    Amper = Label(AggProducto, text='A', width=2,  anchor='w').grid(row=7, column=2)
    Frecuencialbl = Label(AggProducto, text='Frecuencia:', width=11, anchor='w', pady=10).grid(row=8, column=0)
    Claselbl = Label(AggProducto, text='Clase:', width=11, anchor='w', pady=10).grid(row=9, column=0)


    
    botonguardar = Button(AggProducto, text='Guardar', command=lambda: [AgregarProducto(Marca.get(), Modelo.get(), Descripcion.get(), MonoTri.get(), PotenciaN.get(), Tension.get(), CorrienteM.get(), Frecuencia.get(), Clase.get()), ActualizarLista()])
    botonsalir = Button(AggProducto, text='Cancelar', command=lambda:AggProducto.destroy())
    
    Marca.grid(row=1, column=1)
    Modelo.grid(row=2, column=1)
    Descripcion.grid(row=3, column=1)
    MonoTri.grid(row=4, column=1)
    PotenciaN.grid(row=5, column=1)
    Tension.grid(row=6, column=1)
    CorrienteM.grid(row=7, column=1)
    Frecuencia.grid(row=8, column=1)
    Clase.grid(row=9, column=1)
    botonguardar.place(x=60, y=365)
    botonsalir.place(x=150, y=365)

def ActualizarLista():
    listaproductos = []
    for row1 in range(3,2500):
        if(ws1.cell(row1,2).value is not None):
            try:
                MarcaProducto = ws1.cell(row1,2).value + ' - '+ ws1.cell(row1,3).value
                listaproductos.append(MarcaProducto)
            except:
                listaproductos.append(ws1.cell(row1,2).value)
        if(ws1.cell(row1,2).value is None):
            pass
    listaproductos_var = StringVar(value=listaproductos)
    productosbox = Listbox(root, listvariable=listaproductos_var, height=10, selectmode='extended')
    productosbox.bind('<<ListboxSelect>>', CambioClase)
    scrollbar = Scrollbar(root, orient='vertical', command=productosbox.yview, )
    productosbox['yscrollcommand'] = scrollbar.set
    scrollbar.place(x=154, y=30)
    productosbox.place(x=30, y=30)

def delete_text(event):
    global default_text
    if default_text:
        Buscador.delete(0, END)
        default_text = False

barramenu = Menu(root)
root.config(menu=barramenu)

productmenu = Menu(barramenu, tearoff=0)

barramenu.add_cascade(label="Productos", menu=productmenu)

productmenu.add_command(label="Agregar producto", command=VentanaAggProducto)
productmenu.add_command(label="Editar producto")
productmenu.add_command(label="Eliminar producto")

default_text = True

botonon = 'yellow'
botonoff = '#F0F0F0'

separador0 = Label(root, text='  ').grid(row=0, column=0)
separador1 = Label(root, text='                                                            ').grid(row=1, column=0)
separador2 = Label(root, text='  ').grid(row=2, column=3)
separador3 = Label(root, text='  ').grid(row=2, column=5)
separador4 = Label(root, text='  ').grid(row=2, column=7)
separador5 = Label(root, text='  ').grid(row=2, column=9)

boton1 = Button(root, text='Corriente', bg=botonoff, padx='12px', pady='5px', command=lambda: Presionado(boton1))
boton2 = Button(root, text='Potencia', bg=botonoff, padx='12px', pady='5px', command=lambda: Presionado(boton2))
boton3 = Button(root, text="""Rigidez
dieléctrica""", bg=botonoff, padx='12px', pady='0px', command=lambda: Presionado(boton3))
boton4 = Button(root, text="""Puesta
a tierra""", bg=botonoff, padx='12px', pady='0px', command=lambda: Presionado(boton4))
boton5 = Button(root, text="""Corriente 
de fuga""", bg=botonoff, padx='12px', pady='0px', command=lambda: Presionado(boton5))

botonensayo = Button(root, text='Ensayar', bg='green', padx='25px', pady='5px').place(x=300, y=230)

barra1 = ttk.Progressbar(root, style='green.Horizontal.TProgressbar', orient='vertical', value=50, length=150)
barra2 = ttk.Progressbar(root, style='green.Horizontal.TProgressbar', orient='vertical', length=150)
barra3 = ttk.Progressbar(root, style='green.Horizontal.TProgressbar', orient='vertical', length=150)
barra4 = ttk.Progressbar(root, style='green.Horizontal.TProgressbar', orient='vertical', length=150)
barra5 = ttk.Progressbar(root, style='green.Horizontal.TProgressbar', orient='vertical', length=150)

listaproductos = []

for row1 in range(3,2500):
    if(ws1.cell(row1,2).value is not None):
        try:
            MarcaProducto = ws1.cell(row1,2).value + ' - '+ ws1.cell(row1,3).value
            listaproductos.append(MarcaProducto)
        except:
            listaproductos.append(ws1.cell(row1,2).value)
    if(ws1.cell(row1,2).value is None):
        pass

listaproductos_var = StringVar(value=listaproductos)
productosbox = Listbox(root, listvariable=listaproductos_var, height=10, selectmode='extended')

Buscador = Entry(root, bd=3)
Buscador.insert(END, 'Buscador...')
Buscador.bind("<Button-1>", delete_text)
Buscador.place(x=30, y=20)

productosbox.bind('<<ListboxSelect>>', CambioClase)

scrollbar = Scrollbar(root, orient='vertical', command=productosbox.yview,)
productosbox['yscrollcommand'] = scrollbar.set

scrollbar.place(x=154, y=100)
productosbox.place(x=30, y=45)

barra1.grid(row=1, column=2)
boton1.grid(row=2, column=2)
barra2.grid(row=1, column=4)
boton2.grid(row=2, column=4)
barra3.grid(row=1, column=6)
boton3.grid(row=2, column=6)
barra4.grid(row=1, column=8)
boton4.grid(row=2, column=8)
barra5.grid(row=1, column=10)
boton5.grid(row=2, column=10)

root.mainloop()
wb.save('productos.xlsx')
